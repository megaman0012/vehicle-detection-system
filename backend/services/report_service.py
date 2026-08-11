"""
Report Service for Vehicle Detection System
Handles generation of PDF and Excel reports
"""

import io
import os
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional

# For PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# For Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ReportService:
    def __init__(self, db: Optional[Any] = None):
        self.db = db
        self.reports_dir = "reports"
        os.makedirs(self.reports_dir, exist_ok=True)

    def _query_events(self, start: datetime, end: datetime) -> List[Dict[str, Any]]:
        """Query events between two datetimes and return them as dicts for reports"""
        from models.event import Event
        from models.camera import Camera

        rows = (
            self.db.query(Event, Camera.name)
            .outerjoin(Camera, Event.camera_id == Camera.id)
            .filter(Event.timestamp >= start, Event.timestamp < end)
            .all()
        )
        result = []
        for event, camera_name in rows:
            result.append({
                "timestamp": event.timestamp,
                "event_type": event.event_type,
                "description": event.description,
                "camera_name": camera_name or "",
                "license_plate": event.license_plate or "-"
            })
        return result

    def generate_daily_report(self, date) -> List[Dict[str, Any]]:
        """Generate daily report data for the given date"""
        if isinstance(date, str):
            date = datetime.strptime(date, "%Y-%m-%d").date()
        start = datetime.combine(date, datetime.min.time())
        end = start + timedelta(days=1)
        return self._query_events(start, end)

    def generate_weekly_report(self, start_date) -> List[Dict[str, Any]]:
        """Generate weekly report data starting from the given date"""
        if isinstance(start_date, str):
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        start = datetime.combine(start_date, datetime.min.time())
        end = start + timedelta(days=7)
        return self._query_events(start, end)

    def generate_monthly_report(self, year: int, month: int) -> List[Dict[str, Any]]:
        """Generate monthly report data for the given year and month"""
        start = datetime(year, month, 1)
        end = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
        return self._query_events(start, end)

    def generate_pdf_report(self, report_data: List[Dict[str, Any]], period_type: str = "daily") -> io.BytesIO:
        """Generate a PDF report and return it as a BytesIO buffer"""
        return io.BytesIO(self.generate_parking_report_pdf(report_data))

    def generate_excel_report(self, report_data: List[Dict[str, Any]], period_type: str = "daily") -> io.BytesIO:
        """Generate an Excel report and return it as a BytesIO buffer"""
        return io.BytesIO(self.generate_parking_report_excel(report_data))

    def get_statistics(self, days: int = 30) -> Dict[str, Any]:
        """Get system statistics for the last N days"""
        from models.event import Event
        from models.vehicle import DetectedVehicle

        since = datetime.utcnow() - timedelta(days=days)
        total_events = self.db.query(Event).count()
        recent_events = self.db.query(Event).filter(Event.timestamp >= since).count()
        total_vehicles = self.db.query(DetectedVehicle).count()
        parked_vehicles = self.db.query(DetectedVehicle).filter(DetectedVehicle.is_parked.is_(True)).count()

        return {
            "total_events": total_events,
            f"events_last_{days}_days": recent_events,
            "total_vehicles_detected": total_vehicles,
            "currently_parked": parked_vehicles,
            "generated_at": datetime.utcnow().isoformat()
        }
    
    def generate_parking_report_pdf(self, events: List[Dict[str, Any]], 
                                  filters: Dict[str, Any] = None) -> bytes:
        """
        Generate a PDF report of parking events
        
        Args:
            events: List of event dictionaries
            filters: Optional filters applied to the report
            
        Returns:
            PDF file as bytes
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab is required for PDF generation")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph("Reporte de Eventos de Estacionamiento", title_style))
        story.append(Spacer(1, 12))
        
        # Report metadata
        metadata_style = styles['Normal']
        story.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", metadata_style))
        
        if filters:
            filter_text = "Filtros aplicados: "
            filter_parts = []
            if filters.get('event_type'):
                filter_text += f"Tipo: {filters['event_type']}, "
            if filters.get('start_date'):
                filter_text += f"Desde: {filters['start_date']}, "
            if filters.get('end_date'):
                filter_text += f"Hasta: {filters['end_date']}, "
            if filters.get('camera_id'):
                filter_text += f"Cámara: {filters['camera_id']}"
            
            story.append(Paragraph(filter_text.rstrip(', '), metadata_style))
        
        story.append(Spacer(1, 20))
        
        # Events table
        if events:
            # Table header
            data = [['Fecha/Hora', 'Tipo', 'Descripción', 'Cámara', 'Placa']]
            
            # Table rows
            for event in events:
                date_str = event.get('timestamp', '')
                if isinstance(date_str, datetime):
                    date_str = date_str.strftime('%d/%m/%Y %H:%M:%S')
                
                data.append([
                    date_str,
                    event.get('event_type', '').replace('_', ' ').title(),
                    event.get('description', '')[:50] + ('...' if len(event.get('description', '')) > 50 else ''),
                    event.get('camera_name', ''),
                    event.get('license_plate', '-')
                ])
            
            # Create table
            table = Table(data, colWidths=[1.2*inch, 1.2*inch, 2.5*inch, 1.2*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            story.append(table)
        else:
            story.append(Paragraph("No se encontraron eventos para el reporte.", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_parking_report_excel(self, events: List[Dict[str, Any]], 
                                    filters: Dict[str, Any] = None) -> bytes:
        """
        Generate an Excel report of parking events
        
        Args:
            events: List of event dictionaries
            filters: Optional filters applied to the report
            
        Returns:
            Excel file as bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("OpenPyXL is required for Excel generation")
        
        # Create workbook and select active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Eventos de Estacionamiento"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = "Reporte de Eventos de Estacionamiento"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Metadata
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        row_offset = 4
        
        # Filters info
        if filters:
            filter_text = "Filtros aplicados: "
            filter_parts = []
            if filters.get('event_type'):
                filter_text += f"Tipo: {filters['event_type']}, "
            if filters.get('start_date'):
                filter_text += f"Desde: {filters['start_date']}, "
            if filters.get('end_date'):
                filter_text += f"Hasta: {filters['end_date']}, "
            if filters.get('camera_id'):
                filter_text += f"Cámara: {filters['camera_id']}"
            
            ws.merge_cells(f'A{row_offset}:E{row_offset}')
            ws[f'A{row_offset}'] = filter_text.rstrip(', ')
            ws[f'A{row_offset}'].font = Font(italic=True)
            row_offset += 1
        
        # Headers
        headers = ['Fecha/Hora', 'Tipo', 'Descripción', 'Cámara', 'Placa']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_offset, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        row_offset += 1
        
        # Data rows
        for event in events:
            date_str = event.get('timestamp', '')
            if isinstance(date_str, datetime):
                date_str = date_str.strftime('%d/%m/%Y %H:%M:%S')
            
            row_data = [
                date_str,
                event.get('event_type', '').replace('_', ' ').title(),
                event.get('description', ''),
                event.get('camera_name', ''),
                event.get('license_plate', '-')
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_offset, column=col, value=value)
            
            row_offset += 1
        
        # Adjust column widths
        for col in range(1, 6):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_summary_report_pdf(self, stats: Dict[str, Any]) -> bytes:
        """
        Generate a summary PDF report
        
        Args:
            stats: Dictionary containing statistics
            
        Returns:
            PDF file as bytes
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab is required for PDF generation")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph("Reporte Resumen del Sistema", title_style))
        story.append(Spacer(1, 12))
        
        # Report metadata
        metadata_style = styles['Normal']
        story.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", metadata_style))
        story.append(Spacer(1, 20))
        
        # Statistics table
        data = [['Métrica', 'Valor']]
        
        for key, value in stats.items():
            # Format key for display
            display_key = key.replace('_', ' ').title()
            data.append([display_key, str(value)])
        
        # Create table
        table = Table(data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def save_report(self, content: bytes, filename: str) -> str:
        """
        Save report content to file
        
        Args:
            content: Report content as bytes
            filename: Name of the file to save
            
        Returns:
            Path to the saved file
        """
        filepath = os.path.join(self.reports_dir, filename)
        with open(filepath, 'wb') as f:
            f.write(content)
        return filepath

# Singleton instance
report_service = ReportService()